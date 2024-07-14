import os
from openpyxl import Workbook, load_workbook

credentials_path = os.path.join(os.path.expanduser("~"), "Documents", "credenciais")
os.makedirs(credentials_path, exist_ok=True)

class Usuario:
    def __init__(self, email):
        self.email = email
        self.dias_treino = []
        self.experiencia = None
        self.lesao_limitacao = None
        self.treino_selecionado = []
        self.treinos_file = os.path.join(credentials_path, f'{self.email}_treino.xlsx')

    def receber_dias_treino(self, dias_treino):
        self.dias_treino = dias_treino

    def receber_experiencia(self, experiencia):
        self.experiencia = experiencia

    def receber_lesao_limitacao(self, lesao_limitacao):
        self.lesao_limitacao = lesao_limitacao

    def carregar_treinos_excel(self):
        planilha_treinos = os.path.join(os.path.expanduser("~"), "Downloads", 'planilha_treinos.xlsx')
        workbook = load_workbook(filename=planilha_treinos, read_only=True)
        sheet = workbook.active

        treinos = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            grupo_muscular, dificuldade, exercicio1, series1, repeticoes1, exercicio2, series2, repeticoes2, exercicio3, series3, repeticoes3, exercicio4, series4, repeticoes4 = row
            treinos.append({
                'grupo_muscular': grupo_muscular.lower(),
                'dificuldade': dificuldade,
                'exercicios': [
                    {'exercicio': exercicio1, 'series': series1, 'repeticoes': repeticoes1},
                    {'exercicio': exercicio2, 'series': series2, 'repeticoes': repeticoes2},
                    {'exercicio': exercicio3, 'series': series3, 'repeticoes': repeticoes3},
                    {'exercicio': exercicio4, 'series': series4, 'repeticoes': repeticoes4}
                ]
            })

        return treinos

    def selecionar_exercicios(self, grupo_muscular, dificuldade, grupos_selecionados, max_exercicios=3):
        treinos = self.carregar_treinos_excel()
        exercicios_selecionados = []

        for treino in treinos:
            if treino['grupo_muscular'] == grupo_muscular and (str(treino['dificuldade']).lower() == dificuldade.lower()):
                for exercicio in treino['exercicios']:
                    if exercicio['exercicio'] is not None and exercicio['exercicio'] not in grupos_selecionados:
                        exercicios_selecionados.append(exercicio)
                        grupos_selecionados.add(exercicio['exercicio'])

                if len(exercicios_selecionados) >= max_exercicios:
                    break

        return exercicios_selecionados[:max_exercicios]

    def selecionar_e_montar_treino(self, grupos_musculares, max_exercicios, grupos_selecionados):
        treino_dia = []
        for grupo_muscular in grupos_musculares:
            exercicios = self.selecionar_exercicios(grupo_muscular, str(self.experiencia['numero']), grupos_selecionados, max_exercicios=max_exercicios)
            if exercicios:
                treino_dia.append({
                    'grupo_muscular': grupo_muscular.capitalize(),
                    'exercicios': exercicios
                })
        return treino_dia

    def montar_treino(self):
        treino_completo = []
        grupos_por_dia = {
            1: [['peito', 'ombro', 'tríceps', 'costas', 'antebraço', 'bíceps', 'quadríceps', 'posterior de coxa', 'glúteo'], 2],
            2: [['peito', 'ombro', 'tríceps', 'costas', 'antebraço', 'bíceps'], ['quadríceps', 'posterior de coxa', 'glúteo'], 3],
            3: [['peito', 'ombro', 'tríceps'], ['costas', 'antebraço', 'bíceps'], ['quadríceps', 'posterior de coxa', 'glúteo'], 3],
            4: [['peito', 'ombro'], ['costas', 'antebraço'], ['quadríceps', 'posterior de coxa', 'glúteo'], ['bíceps', 'tríceps'], 4],
            5: [['peito'], ['costas'], ['quadríceps', 'posterior de coxa', 'glúteo'], ['bíceps', 'tríceps'], ['ombro', 'antebraço'], 4],
            6: [['peito'], ['costas'], ['quadríceps', 'posterior de coxa', 'glúteo'], ['bíceps'], ['tríceps'], ['ombro', 'antebraço'], 4],
            7: [['peito'], ['costas'], ['quadríceps', 'posterior de coxa', 'glúteo'], ['bíceps'], ['tríceps'], ['ombro'], ['antebraço'], 4]
        }

        num_dias = len(self.dias_treino)
        if num_dias in grupos_por_dia:
            grupos_musculares, max_exercicios = grupos_por_dia[num_dias][:-1], grupos_por_dia[num_dias][-1]
            for i, dia in enumerate(self.dias_treino):
                grupos_selecionados = set()
                treino_dia = self.selecionar_e_montar_treino(grupos_musculares[i], max_exercicios, grupos_selecionados)
                if treino_dia:
                    treino_completo.append({
                        'dia': dia.capitalize(),
                        'treino': treino_dia
                    })
                else:
                    print(f"Não há treino definido para {dia.capitalize()}.")
        else:
            print(f"Número de dias de treino inválido: {num_dias}")

        self.treino_selecionado = treino_completo

    def imprimir_treino(self):
        for treino_dia in self.treino_selecionado:
            print(f"Dia de Treino: {treino_dia['dia']}")
            for grupo in treino_dia['treino']:
                print(f"Grupo Muscular: {grupo['grupo_muscular']}")
                for exercicio in grupo['exercicios']:
                    if exercicio['exercicio'] is not None:
                        print(f"Exercício: {exercicio['exercicio']}, Séries: {exercicio['series']}, Repetições: {exercicio['repeticoes']}")
                print()

    def salvar_treinos_excel(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Treinos"

        sheet.append(['Dia de Treino', 'Grupo Muscular', 'Exercício', 'Séries', 'Repetições'])

        for treino_dia in self.treino_selecionado:
            for grupo in treino_dia['treino']:
                for exercicio in grupo['exercicios']:
                    sheet.append([treino_dia['dia'], grupo['grupo_muscular'], exercicio['exercicio'], exercicio['series'], exercicio['repeticoes']])

        workbook.save(self.treinos_file)
        print(f"Treinos salvos com sucesso em {self.treinos_file}")

    def verificar_treino_salvo(self):
        return os.path.exists(self.treinos_file)

    def carregar_treino_salvo(self):
        if self.verificar_treino_salvo():
            workbook = load_workbook(filename=self.treinos_file, read_only=True)
            sheet = workbook.active

            treino_completo = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                dia, grupo_muscular, exercicio, series, repeticoes = row

                treino_dia = next((td for td in treino_completo if td['dia'] == dia), None)
                if not treino_dia:
                    treino_dia = {'dia': dia, 'treino': []}
                    treino_completo.append(treino_dia)

                grupo = next((g for g in treino_dia['treino'] if g['grupo_muscular'] == grupo_muscular), None)
                if not grupo:
                    grupo = {'grupo_muscular': grupo_muscular, 'exercicios': []}
                    treino_dia['treino'].append(grupo)

                grupo['exercicios'].append({'exercicio': exercicio, 'series': series, 'repeticoes': repeticoes})

            self.treino_selecionado = treino_completo
