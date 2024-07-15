import os
import tkinter as tk
from tkinter import messagebox
from tkinter import ttk
from openpyxl import Workbook, load_workbook
from usuario import Usuario

# Defina o caminho para a pasta de credenciais e treinos
credentials_path = os.path.join(os.path.expanduser("~"), "Documents", "credenciais")
os.makedirs(credentials_path, exist_ok=True)
credentials_file = os.path.join(credentials_path, 'credentials.xlsx')


class LoginSystem:
    def __init__(self, root):
        """
        Inicializa a interface gráfica do sistema de login.

        Parâmetros:
        root (tk.Tk): A janela principal da aplicação Tkinter.
        """
        self.root = root
        self.root.title("Sistema de Login")

        # Inicializa o Excel com cabeçalhos se necessário
        self.initialize_excel()

        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(pady=20, padx=20)

        ttk.Button(main_frame, text="Registrar", command=self.open_register_window).pack(pady=10)
        ttk.Button(main_frame, text="Login", command=self.open_login_window).pack(pady=10)

    def initialize_excel(self):
        """
        Inicializa o arquivo Excel para armazenar credenciais, adicionando cabeçalhos se necessário.
        """
        if not os.path.exists(credentials_file):
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Credenciais"
            sheet.append(['Name', 'Email', 'Password'])
            workbook.save(credentials_file)

    def save_credentials(self, name, email, password):
        """
        Salva as credenciais do usuário no arquivo Excel.

        Parâmetros:
        name (str): Nome do usuário.
        email (str): Email do usuário.
        password (str): Senha do usuário.

        Retorna:
        bool: True se as credenciais foram salvas com sucesso, False caso contrário.
        """
        try:
            workbook = load_workbook(credentials_file)
            sheet = workbook.active
            sheet.append([name, email, password])
            workbook.save(credentials_file)
            return True
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao salvar credenciais: {e}")
            return False

    def verify_credentials(self, email, password):
        """
        Verifica as credenciais do usuário no arquivo Excel.

        Parâmetros:
        email (str): Email do usuário.
        password (str): Senha do usuário.

        Retorna:
        bool: True se as credenciais forem válidas, False caso contrário.
        """
        try:
            workbook = load_workbook(credentials_file, read_only=True)
            sheet = workbook.active
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[1] == email and row[2] == password:
                    return True
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao verificar credenciais: {e}")
        return False

    def register_user(self, register_window, name, email, password):
        """
        Registra um novo usuário após verificar se todos os campos foram preenchidos.

        Parâmetros:
        register_window (tk.Toplevel): A janela de registro.
        name (str): Nome do usuário.
        email (str): Email do usuário.
        password (str): Senha do usuário.
        """
        if name and email and password:
            if self.save_credentials(name, email, password):
                messagebox.showinfo("Sucesso", "Usuário registrado com sucesso!")
                register_window.destroy()
        else:
            messagebox.showwarning("Aviso", "Por favor, preencha todos os campos")

    def login_user(self, email, password):
        """
        Realiza o login do usuário após verificar as credenciais.

        Parâmetros:
        email (str): Email do usuário.
        password (str): Senha do usuário.
        """
        if email and password:
            if self.verify_credentials(email, password):
                messagebox.showinfo("Sucesso", "Login realizado com sucesso!")
                self.root.destroy()
                from interface_treino import InterfaceTreino  # Import aqui para evitar importações circulares
                usuario = Usuario(email)
                interface_treino = InterfaceTreino(usuario)
                interface_treino.exibir_interface()
            else:
                messagebox.showwarning("Erro", "Credenciais inválidas")
        else:
            messagebox.showwarning("Aviso", "Por favor, preencha todos os campos")

    def open_register_window(self):
        """
        Abre a janela de registro para o usuário inserir seus dados.
        """
        register_window = tk.Toplevel(self.root)
        register_window.title("Registrar")

        frame = ttk.Frame(register_window, padding="10")
        frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        ttk.Label(frame, text="Nome:").grid(row=0, column=0, padx=10, pady=5, sticky=tk.W)
        name_entry = ttk.Entry(frame, width=30)
        name_entry.grid(row=0, column=1, padx=10, pady=5, sticky=tk.E)

        ttk.Label(frame, text="E-mail:").grid(row=1, column=0, padx=10, pady=5, sticky=tk.W)
        email_entry = ttk.Entry(frame, width=30)
        email_entry.grid(row=1, column=1, padx=10, pady=5, sticky=tk.E)

        ttk.Label(frame, text="Senha:").grid(row=2, column=0, padx=10, pady=5, sticky=tk.W)
        password_entry = ttk.Entry(frame, width=30, show="*")
        password_entry.grid(row=2, column=1, padx=10, pady=5, sticky=tk.E)

        register_button = ttk.Button(frame, text="Registrar", command=lambda: self.register_user(
            register_window, name_entry.get(), email_entry.get(), password_entry.get()
        ))
        register_button.grid(row=3, columnspan=2, pady=10)

    def open_login_window(self):
        """
        Abre a janela de login para o usuário inserir suas credenciais.
        """
        login_window = tk.Toplevel(self.root)
        login_window.title("Login")

        frame = ttk.Frame(login_window, padding="10")
        frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        ttk.Label(frame, text="E-mail:").grid(row=0, column=0, padx=10, pady=5, sticky=tk.W)
        email_entry = ttk.Entry(frame, width=30)
        email_entry.grid(row=0, column=1, padx=10, pady=5, sticky=tk.E)

        ttk.Label(frame, text="Senha:").grid(row=1, column=0, padx=10, pady=5, sticky=tk.W)
        password_entry = ttk.Entry(frame, width=30, show="*")
        password_entry.grid(row=1, column=1, padx=10, pady=5, sticky=tk.E)

        login_button = ttk.Button(frame, text="Login", command=lambda: self.login_user(
            email_entry.get(), password_entry.get()
        ))
        login_button.grid(row=2, columnspan=2, pady=10)


# Iniciar o sistema de login
if __name__ == "__main__":
    root = tk.Tk()
    app = LoginSystem(root)
    root.mainloop()
