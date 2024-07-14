from openpyxl import load_workbook, Workbook
import os

class Usuario:
    def __init__(self):
        self.dias_treino = []   #Variavel para armazenar os dias que a pessoa treina
        self.experiencia = None  #Variavel que armazena  experiencia do usuario
        self.lesao_limitacao = None #Variavel que armazena a lesão
        self.treino_selecionado = []  #Variavel que armazena o treino

    def receber_dias_treino(self, dias_treino):
        self.dias_treino = dias_treino

    def receber_experiencia(self, experiencia):
        self.experiencia = experiencia

    def receber_lesao_limitacao(self, lesao_limitacao):
        self.lesao_limitacao = lesao_limitacao


    def carregar_treinos_excel(self):
        caminho_downloads = os.path.join(os.path.expanduser("~"), "Downloads")    #Acessa a pasta dowloads
        arquivo_excel = os.path.join(caminho_downloads, 'planilha_treinos.xlsx')   #Procura a pasta com os treinos

        workbook = load_workbook(filename=arquivo_excel, read_only=True)
        sheet = workbook.active

        treinos = [] # Inicializa a lista de treinos

        for row in sheet.iter_rows(min_row=2, values_only=True): # Itera pelas linhas da planilha
            grupo_muscular, dificuldade, exercicio1, series1, repeticoes1, exercicio2, series2, repeticoes2, exercicio3, series3, repeticoes3, exercicio4, series4, repeticoes4 = row # Extrai e organiza os dados de cada linha
            treinos.append({              # Armazena os dados em um dicionário e adiciona à lista de treinos
                'grupo_muscular': grupo_muscular.lower(),
                'dificuldade': dificuldade,
                'exercicios': [
                    {'exercicio': exercicio1, 'series': series1, 'repeticoes': repeticoes1},
                    {'exercicio': exercicio2, 'series': series2, 'repeticoes': repeticoes2},
                    {'exercicio': exercicio3, 'series': series3, 'repeticoes': repeticoes3},
                    {'exercicio': exercicio4, 'series': series4, 'repeticoes': repeticoes4}
                ]
            })

        return treinos

    def selecionar_exercicios(self, grupo_muscular, dificuldade, grupos_selecionados, max_exercicios=3):
        treinos = self.carregar_treinos_excel() # Carregar treinos do Excel
        exercicios_selecionados = [] # Inicializar a lista de exercícios selecionados

        for treino in treinos: # Itera pelos treinos carregados
            if treino['grupo_muscular'] == grupo_muscular and (str(treino['dificuldade']).lower() == dificuldade.lower() ):
                for exercicio in treino['exercicios']: # Selecionar exercícios não duplicados
                    if exercicio['exercicio'] is not None and exercicio['exercicio'] not in grupos_selecionados:
                        exercicios_selecionados.append(exercicio)
                        grupos_selecionados.add(exercicio['exercicio'])

                if len(exercicios_selecionados) >= max_exercicios: # Verifica o limite de exercícios selecionados
                    break

        return exercicios_selecionados[:max_exercicios]  # Retorna a lista limitada de exercícios

    def montar_treino(self):
        treino_completo = [] # Inicializar a lista de treinos completos
        if len(self.dias_treino) == 1: # Verificar se há apenas um dia de treino
            treino_dia = [] # Treino de um dia na semana
            grupos_selecionados = set()
            for grupo_muscular in ['peito', 'ombro', 'tríceps', 'costas', 'antebraço', 'bíceps', 'quadríceps', 'posterior de coxa', 'glúteo']: #Iterar pelos grupos musculares e selecionar exercícios
                exercicios = self.selecionar_exercicios(grupo_muscular, str(self.experiencia['numero']), grupos_selecionados, max_exercicios=2)
                if exercicios:
                    treino_dia.append({
                        'grupo_muscular': grupo_muscular.capitalize(),
                        'exercicios': exercicios
                    })
            if treino_dia: # Adicionar o treino do dia ao treino completo
                treino_completo.append({
                    'dia': self.dias_treino[0].capitalize(),
                    'treino': treino_dia
                })
            else:
                print(f"Não há treino definido para {self.dias_treino[0].capitalize()}.")

        elif len(self.dias_treino) == 2: # Verificar se a pessoa colcoou dois dias de treino
            for i, dia in enumerate(self.dias_treino):
                treino_dia = [] # Treino de um dia na semana
                grupos_selecionados = set()
                if i == 0:  # Primeiro dia: Superiores
                    for grupo_muscular in ['peito', 'ombro', 'tríceps', 'costas', 'antebraço', 'bíceps']:
                        exercicios = self.selecionar_exercicios(grupo_muscular, str(self.experiencia['numero']), grupos_selecionados, max_exercicios=3)
                        if exercicios:
                            treino_dia.append({
                                'grupo_muscular': grupo_muscular.capitalize(),
                                'exercicios': exercicios
                            })
                elif i == 1:  # Segundo dia: Inferiores
                    for grupo_muscular in ['quadríceps', 'posterior de coxa', 'glúteo']:
                        exercicios = self.selecionar_exercicios(grupo_muscular, str(self.experiencia['numero']), grupos_selecionados, max_exercicios=3)
                        if exercicios:
                            treino_dia.append({
                                'grupo_muscular': grupo_muscular.capitalize(),
                                'exercicios': exercicios
                            })

                if treino_dia: # Adicionar o treino do dia ao treino completo
                    treino_completo.append({
                        'dia': dia.capitalize(),
                        'treino': treino_dia
                    })
                else:
                    print(f"Não há treino definido para {dia.capitalize()}.")

        elif len(self.dias_treino) == 3: # Verificar se a pessoa colcoou 3 dias de treino
            # Treino dividido em três dias na semana
            for i, dia in enumerate(self.dias_treino):
                treino_dia = []
                grupos_selecionados = set()
                if i == 0:  # Primeiro dia: Peito, Ombro, Tríceps
                    for grupo_muscular in ['peito', 'ombro', 'tríceps']:
                        exercicios = self.selecionar_exercicios(grupo_muscular, str(self.experiencia['numero']), grupos_selecionados, max_exercicios=3)
                        if exercicios:
                            treino_dia.append({
                                'grupo_muscular': grupo_muscular.capitalize(),
                                'exercicios': exercicios
                            })
                elif i == 1:  # Segundo dia: Costas, Antebraço, Bíceps
                    for grupo_muscular in ['costas', 'antebraço', 'bíceps']:
                        exercicios = self.selecionar_exercicios(grupo_muscular, str(self.experiencia['numero']), grupos_selecionados, max_exercicios=3)
                        if exercicios:
                            treino_dia.append({
                                'grupo_muscular': grupo_muscular.capitalize(),
                                'exercicios': exercicios
                            })
                elif i == 2:  # Terceiro dia: Quadríceps, Posterior de coxa, Glúteo
                    for grupo_muscular in ['quadríceps', 'posterior de coxa', 'glúteo']:
                        exercicios = self.selecionar_exercicios(grupo_muscular, str(self.experiencia['numero']), grupos_selecionados, max_exercicios=3)
                        if exercicios:
                            treino_dia.append({
                                'grupo_muscular': grupo_muscular.capitalize(),
                                'exercicios': exercicios
                            })
                if treino_dia: # Adicionar o treino do dia ao treino completo
                    treino_completo.append({
                        'dia': dia.capitalize(),
                        'treino': treino_dia
                    })
                else:
                    print(f"Não há treino definido para {dia.capitalize()}.")

        elif len(self.dias_treino) == 4:
            # Treino dividido em três dias na semana
            for i, dia in enumerate(self.dias_treino):
                treino_dia = []
                grupos_selecionados = set()
                if i == 0:  # Primeiro dia: Peito, Ombro
                    for grupo_muscular in ['peito', 'ombro']:
                        exercicios = self.selecionar_exercicios(grupo_muscular, str(self.experiencia['numero']),
                                                                grupos_selecionados, max_exercicios=4)
                        if exercicios:
                            treino_dia.append({
                                'grupo_muscular': grupo_muscular.capitalize(),
                                'exercicios': exercicios
                            })
                elif i == 1:  # Segundo dia: Costas, Antebraço,
                    for grupo_muscular in ['costas', 'antebraço']:
                        exercicios = self.selecionar_exercicios(grupo_muscular, str(self.experiencia['numero']),
                                                                grupos_selecionados, max_exercicios=4)
                        if exercicios:
                            treino_dia.append({
                                'grupo_muscular': grupo_muscular.capitalize(),
                                'exercicios': exercicios
                            })
                elif i == 2:  # Terceiro dia: Quadríceps, Posterior de coxa, Glúteo
                    for grupo_muscular in ['quadríceps', 'posterior de coxa', 'glúteo']:
                        exercicios = self.selecionar_exercicios(grupo_muscular, str(self.experiencia['numero']),
                                                                grupos_selecionados, max_exercicios=4)
                        if exercicios:
                            treino_dia.append({
                                'grupo_muscular': grupo_muscular.capitalize(),
                                'exercicios': exercicios
                            })
                elif i == 3:  # Quarto dia : Biceps e triceps
                    for grupo_muscular in ['bíceps', 'tríceps']:
                        exercicios = self.selecionar_exercicios(grupo_muscular, str(self.experiencia['numero']),
                                                                grupos_selecionados, max_exercicios=4)
                        if exercicios:
                            treino_dia.append({
                                'grupo_muscular': grupo_muscular.capitalize(),
                                'exercicios': exercicios
                            })
                if treino_dia: # Adicionar o treino do dia ao treino completo
                    treino_completo.append({
                        'dia': dia.capitalize(),
                        'treino': treino_dia
                    })
                else:
                    print(f"Não há treino definido para {dia.capitalize()}.")
        elif len(self.dias_treino) == 5:
            # Treino dividido em três dias na semana
            for i, dia in enumerate(self.dias_treino):
                treino_dia = []
                grupos_selecionados = set()
                if i == 0:  # Primeiro dia: Peito
                    for grupo_muscular in ['peito']:
                        exercicios = self.selecionar_exercicios(grupo_muscular, str(self.experiencia['numero']),
                                                                grupos_selecionados, max_exercicios=4)
                        if exercicios:
                            treino_dia.append({
                                'grupo_muscular': grupo_muscular.capitalize(),
                                'exercicios': exercicios
                            })
                elif i == 1:  # Segundo dia: Costas
                    for grupo_muscular in ['costas']:
                        exercicios = self.selecionar_exercicios(grupo_muscular, str(self.experiencia['numero']),
                                                                grupos_selecionados, max_exercicios=4)
                        if exercicios:
                            treino_dia.append({
                                'grupo_muscular': grupo_muscular.capitalize(),
                                'exercicios': exercicios
                            })
                elif i == 2:  # Terceiro dia: Quadríceps, Posterior de coxa, Glúteo
                    for grupo_muscular in ['quadríceps', 'posterior de coxa', 'glúteo']:
                        exercicios = self.selecionar_exercicios(grupo_muscular, str(self.experiencia['numero']),
                                                                grupos_selecionados, max_exercicios=4)
                        if exercicios:
                            treino_dia.append({
                                'grupo_muscular': grupo_muscular.capitalize(),
                                'exercicios': exercicios
                            })
                elif i == 3:  # Quarto dia : Biceps e triceps
                    for grupo_muscular in ['bíceps', 'tríceps']:
                        exercicios = self.selecionar_exercicios(grupo_muscular, str(self.experiencia['numero']),
                                                                grupos_selecionados, max_exercicios=4)
                        if exercicios:
                            treino_dia.append({
                                'grupo_muscular': grupo_muscular.capitalize(),
                                'exercicios': exercicios
                            })
                elif i == 4:  # Quinto : Ombro e Antebraçoo
                    for grupo_muscular in ['ombro', 'antebraço']:
                        exercicios = self.selecionar_exercicios(grupo_muscular, str(self.experiencia['numero']),
                                                                grupos_selecionados, max_exercicios=4)
                        if exercicios:
                            treino_dia.append({
                                'grupo_muscular': grupo_muscular.capitalize(),
                                'exercicios': exercicios
                            })
                if treino_dia: # Adicionar o treino do dia ao treino completo
                    treino_completo.append({
                        'dia': dia.capitalize(),
                        'treino': treino_dia
                    })
                else:
                    print(f"Não há treino definido para {dia.capitalize()}.")
        elif len(self.dias_treino) == 6:
            # Treino dividido em três dias na semana
            for i, dia in enumerate(self.dias_treino):
                treino_dia = []
                grupos_selecionados = set()
                if i == 0:  # Primeiro dia: Peito
                    for grupo_muscular in ['peito']:
                        exercicios = self.selecionar_exercicios(grupo_muscular, str(self.experiencia['numero']),
                                                                grupos_selecionados, max_exercicios=4)
                        if exercicios:
                            treino_dia.append({
                                'grupo_muscular': grupo_muscular.capitalize(),
                                'exercicios': exercicios
                            })
                elif i == 1:  # Segundo dia: Costas
                    for grupo_muscular in ['costas']:
                        exercicios = self.selecionar_exercicios(grupo_muscular, str(self.experiencia['numero']),
                                                                grupos_selecionados, max_exercicios=4)
                        if exercicios:
                            treino_dia.append({
                                'grupo_muscular': grupo_muscular.capitalize(),
                                'exercicios': exercicios
                            })
                elif i == 2:  # Terceiro dia: Quadríceps, Posterior de coxa, Glúteo
                    for grupo_muscular in ['quadríceps', 'posterior de coxa', 'glúteo']:
                        exercicios = self.selecionar_exercicios(grupo_muscular, str(self.experiencia['numero']),
                                                                grupos_selecionados, max_exercicios=4)
                        if exercicios:
                            treino_dia.append({
                                'grupo_muscular': grupo_muscular.capitalize(),
                                'exercicios': exercicios
                            })
                elif i == 3:  # Quarto dia : Biceps
                    for grupo_muscular in ['bíceps']:
                        exercicios = self.selecionar_exercicios(grupo_muscular, str(self.experiencia['numero']),
                                                                grupos_selecionados, max_exercicios=4)
                        if exercicios:
                            treino_dia.append({
                                'grupo_muscular': grupo_muscular.capitalize(),
                                'exercicios': exercicios
                            })
                elif i == 4:  # Quinto dia : triceps
                    for grupo_muscular in ['tríceps']:
                        exercicios = self.selecionar_exercicios(grupo_muscular, str(self.experiencia['numero']),
                                                                grupos_selecionados, max_exercicios=4)
                        if exercicios:
                            treino_dia.append({
                                'grupo_muscular': grupo_muscular.capitalize(),
                                'exercicios': exercicios
                            })
                elif i == 5:  # Sexto dia : Ombro e Antebraçoo
                    for grupo_muscular in ['ombro', 'antebraço']:
                        exercicios = self.selecionar_exercicios(grupo_muscular, str(self.experiencia['numero']),
                                                                grupos_selecionados, max_exercicios=4)
                        if exercicios:
                            treino_dia.append({
                                'grupo_muscular': grupo_muscular.capitalize(),
                                'exercicios': exercicios
                            })
                if treino_dia: # Adicionar o treino do dia ao treino completo
                    treino_completo.append({
                        'dia': dia.capitalize(),
                        'treino': treino_dia
                    })
                else:
                    print(f"Não há treino definido para {dia.capitalize()}.")
        elif len(self.dias_treino) == 7:
            # Treino dividido em três dias na semana
            for i, dia in enumerate(self.dias_treino):
                treino_dia = []
                grupos_selecionados = set()
                if i == 0:  # Primeiro dia: Peito
                    for grupo_muscular in ['peito']:
                        exercicios = self.selecionar_exercicios(grupo_muscular, str(self.experiencia['numero']),
                                                                grupos_selecionados, max_exercicios=4)
                        if exercicios:
                            treino_dia.append({
                                'grupo_muscular': grupo_muscular.capitalize(),
                                'exercicios': exercicios
                            })
                elif i == 1:  # Segundo dia: Costas
                    for grupo_muscular in ['costas']:
                        exercicios = self.selecionar_exercicios(grupo_muscular, str(self.experiencia['numero']),
                                                                grupos_selecionados, max_exercicios=4)
                        if exercicios:
                            treino_dia.append({
                                'grupo_muscular': grupo_muscular.capitalize(),
                                'exercicios': exercicios
                            })
                elif i == 2:  # Terceiro dia: Quadríceps, Posterior de coxa, Glúteo
                    for grupo_muscular in ['quadríceps', 'posterior de coxa', 'glúteo']:
                        exercicios = self.selecionar_exercicios(grupo_muscular, str(self.experiencia['numero']),
                                                                grupos_selecionados, max_exercicios=4)
                        if exercicios:
                            treino_dia.append({
                                'grupo_muscular': grupo_muscular.capitalize(),
                                'exercicios': exercicios
                            })
                elif i == 3:  # Quarto dia : Biceps
                    for grupo_muscular in ['bíceps']:
                        exercicios = self.selecionar_exercicios(grupo_muscular, str(self.experiencia['numero']),
                                                                grupos_selecionados, max_exercicios=4)
                        if exercicios:
                            treino_dia.append({
                                'grupo_muscular': grupo_muscular.capitalize(),
                                'exercicios': exercicios
                            })
                elif i == 4:  # Quinto dia : triceps
                    for grupo_muscular in ['tríceps']:
                        exercicios = self.selecionar_exercicios(grupo_muscular, str(self.experiencia['numero']),
                                                                grupos_selecionados, max_exercicios=4)
                        if exercicios:
                            treino_dia.append({
                                'grupo_muscular': grupo_muscular.capitalize(),
                                'exercicios': exercicios
                            })
                elif i == 5:  # Sexto dia : Ombro
                    for grupo_muscular in ['ombro']:
                        exercicios = self.selecionar_exercicios(grupo_muscular, str(self.experiencia['numero']),
                                                                grupos_selecionados, max_exercicios=4)
                        if exercicios:
                            treino_dia.append({
                                'grupo_muscular': grupo_muscular.capitalize(),
                                'exercicios': exercicios
                            })
                elif i == 6:  # Setimo dia : Antebraço
                    for grupo_muscular in ['antebraço']:
                        exercicios = self.selecionar_exercicios(grupo_muscular, str(self.experiencia['numero']),
                                                                grupos_selecionados, max_exercicios=4)
                        if exercicios:
                            treino_dia.append({
                                'grupo_muscular': grupo_muscular.capitalize(),
                                'exercicios': exercicios
                            })

                if treino_dia:
                    treino_completo.append({
                        'dia': dia.capitalize(),
                        'treino': treino_dia
                    })
                else:
                    print(f"Não há treino definido para {dia.capitalize()}.")

        self.treino_selecionado = treino_completo

    def imprimir_treino(self):
        for treino_dia in self.treino_selecionado:
            print(f"Dia de Treino: {treino_dia['dia']}")
            for grupo in treino_dia['treino']:
                print(f"Grupo Muscular: {grupo['grupo_muscular']}")
                for exercicio in grupo['exercicios']:
                    if exercicio['exercicio'] is not None:  # Verifica se há exercício definido
                        print(f"Exercício: {exercicio['exercicio']}, Séries: {exercicio['series']}, Repetições: {exercicio['repeticoes']}")
                print()  # Linha em branco entre os grupos musculares

    def salvar_treinos_excel(self):
        caminho_downloads = os.path.join(os.path.expanduser("~"), "Downloads") # Acessa a pasta Downloads
        arquivo_excel = os.path.join(caminho_downloads, 'treinos_usuario.xlsx') # Defini o caminho do arquivo Excel

        workbook = Workbook() # Cria um novo arquivo Excel
        sheet = workbook.active
        sheet.title = "Treinos"


        sheet.append(['Dia de Treino', 'Grupo Muscular', 'Exercício', 'Séries', 'Repetições']) # Escrever cabeçalhos


        for treino_dia in self.treino_selecionado:  # Escrever dados dos treinos
            for grupo in treino_dia['treino']:
                for exercicio in grupo['exercicios']:
                    sheet.append(
                        [treino_dia['dia'], grupo['grupo_muscular'], exercicio['exercicio'], exercicio['series'],
                         exercicio['repeticoes']])

        workbook.save(filename=arquivo_excel) # Salva o arquivo Excel
        print(f"Treinos salvos com sucesso em {arquivo_excel}")